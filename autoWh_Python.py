from openpyxl import load_workbook
from openpyxl import Workbook
import openpyxl
import numpy
import pandas as pd

#加载参数文件
data1 = pd.read_csv('cn_5934_muestras_todos_los_meses.csv', encoding='utf-8')
data1["Production"] = data1["Production"] * 1000
data1.to_excel('cn_5934_muestras_todos_los_meses.xlsx', sheet_name='Production')

wb1 = load_workbook('cn_5934_muestras_todos_los_meses.xlsx')
wb2 = load_workbook('Comparativa_Susana.xlsx')


#读取workbook中所有表格
sheets1 = wb1.get_sheet_names()
sheets2 = wb2.get_sheet_names()
#打印所有表的名字
print(sheets1)
print(sheets2)

#遍历每个sheet的数据
sheet1 = wb1.get_sheet_by_name(sheets1[0])
sheet2 = wb2.get_sheet_by_name(sheets2[0])

max_row = sheet2.max_row                    #最大行数
max_column = sheet2.max_column              #最大列数

#填写4-1到12-31数据
for m in range(2, max_row+1-2160):
    n = 101                             #chr(97)='a'
    o = 108
    n = chr(n)                          #ASCII字符
    o = chr(o)
    i = '%s%d' % (o, m+2160)                 #读取单元格编号
    j = '%s%d' % (n, m)                 #写入单元格编号
    cell1 = sheet1[i].value             #获取data单元格数据
    sheet2[j].value = cell1             #赋值到test单元格

#填写1-1到2-28数据
for m in range(6602, 8018):
    n = 101                             #chr(97)='a'
    o = 108
    n = chr(n)                          #ASCII字符
    o = chr(o)
    i = '%s%d' % (o, m-6600)                 #读取单元格编号
    j = '%s%d' % (n, m)                 #写入单元格编号
    cell1 = sheet1[i].value             #获取data单元格数据
    sheet2[j].value = cell1             #赋值到test单元格

#填写2-29数据
for m in range(8018, 8042):
    n = 101                             #chr(97)='a'
    o = 108
    n = chr(n)                          #ASCII字符
    o = chr(o)
    i = '%s%d' % (o, m-6624)            #读取单元格编号
    j = '%s%d' % (n, m)                 #写入单元格编号
    cell1 = sheet1[i].value             #获取data单元格数据
    sheet2[j].value = cell1             #赋值到test单元格

#填写3-1到3-31数据
for m in range(8042, max_row+1):
    n = 101                             #chr(97)='a'
    o = 108
    n = chr(n)                          #ASCII字符
    o = chr(o)
    i = '%s%d' % (o, m-6624)            #读取单元格编号
    j = '%s%d' % (n, m)                 #写入单元格编号
    cell1 = sheet1[i].value             #获取data单元格数据
    sheet2[j].value = cell1             #赋值到test单元格

wb2.save('Comparativa_Susana-test.xlsx')    #保存数据
print("写入完成")
wb1.close()                                 #关闭excel
wb2.close()
