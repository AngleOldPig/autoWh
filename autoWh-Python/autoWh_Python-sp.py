from openpyxl import load_workbook
import pandas as pd
import csv
import os
import openpyxl
import re
import numpy as np
np.set_printoptions(threshold=np.inf)

# 需要填写的参数：
# *** 注意 *** 不要填写后缀名，也不要填写文件名的日期部分
predictFileName = '5934_muestras_todos_los_meses'  # 要读取的预测数据文档名
actualFileName = 'Balance_energético'  # 要读取的实际数据文档名
writeFileName = 'Comparativa_Susana'  # 需要录入的表格名
resultFileName = 'Comparativa_Susana-finish'  # 保存结果的文档名
# *** 注意 *** 不要填写后缀名，也不要填写文件名的日期部分


# 加载参数文件：
print("读取中。。。")

# 使用pd.read_csv需加入参数
# 国内（delimiter=";", decimal=",")
# 西班牙（delimiter=";"）


# 加工预测数据文档
data1 = pd.read_csv(predictFileName + '.csv', delimiter=";", thousands='.', encoding='utf-8')
data1["Production"] = data1["Production"] * 1000
data1.to_excel('somethingYouNeed.xlsx', sheet_name='Prediction')

# 加工实际数据文档：

# 生成空模板csv文档，并添加年、月、日、小时、每日产电量，5列标题
actualProductionCsvName = "somethingYouNeedToo.csv"
with open(actualProductionCsvName, 'w') as csvfile:
    csv_writer = csv.writer(csvfile)
    csv_head = ["year", "month", "day", "hour", "production"]
    csv_writer.writerow(csv_head)

# 自动生成真实日期数字
for y in range(2019, 2021):
    for m in range(1, 13):
        for d in range(1, 32):
            if m == 2 and d > 29:
                continue
            if m == 4 or m == 6 or m == 9 or m == 11:
                if d > 30:
                    continue
            # 使一位数字前面加0变成两位数字
            year = str(y)
            month = str(m)
            if m < 10:
                month = '0' + month
            day = str(d)
            if d < 10:
                day = '0' + day
            # 组合生成文件名
            data2FileName = actualFileName + '_' + year + '_' + month + '_' + day + '.csv'
            # 判断当前名称的文件是否存在
            if os.path.exists(data2FileName):
                if os.path.getsize(data2FileName):
                    print('已读取' + data2FileName)
                else:
                    print(data2FileName + '文件存在但为空')
                    continue
            else:
                # print(data2FileName + '文件不存在')
                continue
            # 用暂存变量dataTemp读取生成文档名指向的csv文件
            dataTemp = pd.read_csv(data2FileName, delimiter=";", thousands='.', encoding='utf-8', header=None, skiprows=1,
                                   usecols=[0, 8])
            # 去除数据中的. 防止数据被识别为小数

            # 准备参数
            dataTempList = []
            a = 0
            h = 0
            num = [0.0, 0.0, 0.0, 0.0, 0.0]

            # 调试用的信息
            # print('表格模板：')
            # print(dataTempList)
            # print('选取项：')
            # print(dataTemp[0][3])
            # print('所有项：')
            # print(dataTemp)

            # 将文件中的实际值读取出并和平均值一起存入num数组
            for i in range(0, 96):
                p = dataTemp[8][i]
                num[i % 4] = float(p)                               # 读取每小时的4个数据单元格
                if (i + 1) % 4 == 0:
                    num[4] = num[0] + num[1] + num[2] + num[3]
                    num[4] = num[4] / 4                             # 求它们的平均数
                    a = a + 1
                    dataTempList.append([y,m,d,a,num[4]])           # 将时间数据和每小时平均数添加进dataTempList列表

            # 将 列表list 转换为 DataFrame格式
            data2TempDataFrame = pd.DataFrame(dataTempList)
            # print('表格雏形：')
            # print(data2TempDataFrame)

            # 将时间信息与每小时平均数一并存入csv文件：
            data2TempDataFrame.to_csv(actualProductionCsvName, mode='a', header=False, index=None)

            '''
            # 使用官方csv库尝试失败
            # dataTemp = pd.read_csv(data2FileName, delimiter=";", decimal=",", encoding='utf-8', header=None)
            # 计算每日平均数并存入csv文件N列
            # a = 1
            # m = 0
            # num = [0.0, 0.0, 0.0, 0.0, 0.0]
            # with open(data2FileName, encoding='utf-8') as dataTemp:
            #     # ignore = dataTemp.readline()                            #忽略标题行
            #     for line in csv.DictReader(dataTemp):
            #         num[m] = float(line['Generación fotovoltaica / Promedios [W]  '])
            #         if a % 4 == 0:
            #             num[4] = num[0]+num[1]+num[2]+num[3]
            #             num[4] = num[4]/4
            #             print(line)
            #             m = -1
            #         a = a + 1
            #         m = m + 1
            '''

            # 将年月日小时数据分别放在J,K,L,M列
            # 将年月日小时和平均数5列占据的所有行附加到data2现有数据下面
            # with open(actualProductionCsvName, 'a+') as csvfile:
            #     csv_writer = csv.writer(csvfile)
            #     data_row = ["test"]
            #     csv_writer.writerow(data_row)

# data2读取实际数据csv
data2 = pd.read_csv(actualProductionCsvName, encoding='utf-8')
data2.to_excel('somethingYouNeedToo.xlsx', sheet_name='Production')
data2.to_excel(actualFileName+'-actualProduction.xlsx', sheet_name='Production')

# 读取需要录入的表格
wb1 = load_workbook('somethingYouNeed.xlsx')
wb3 = load_workbook(writeFileName + '.xlsx')

# 读取workbook中所有表格
print('请忽略以下4条报警：')
sheets1 = wb1.get_sheet_names()
sheets3 = wb3.get_sheet_names()

# 遍历每个sheet的数据
sheet1 = wb1.get_sheet_by_name(sheets1[0])
sheet3 = wb3.get_sheet_by_name(sheets3[0])

max_row = sheet3.max_row  # 最大行数
max_column = sheet3.max_column  # 最大列数


# 构建数据填写函数
# 用法：
#
def xlsxDataCopy(columnRead, columnWrite, rowWriteStart, rowWriteStop, rowReadFixed):
    "根据行列编号复制单元格内容"
    for m in range(rowWriteStart, rowWriteStop):
        n = columnWrite
        o = columnRead
        n = chr(n)
        o = chr(o)
        i = '%s%d' % (o, m + rowReadFixed)  # 读取单元格编号
        j = '%s%d' % (n, m)  # 写入单元格编号
        cell1 = sheet1[i].value  # 获取data单元格数据
        sheet3[j].value = cell1  # 赋值到test单元格


# 写入数据：
print("写入中。。。")

# 写入预测数据
xlsxDataCopy(108, 101, 2, max_row + 1 - 2160, 2160)
xlsxDataCopy(108, 101, 6602, 8018, -6600)
xlsxDataCopy(108, 101, 8018, 8042, -6624)
xlsxDataCopy(108, 101, 8042, max_row + 1, -6624)

# 写入实际数据


# 写入预测数据旧代码，现已使用函数代替
# # 填写4-1到12-31数据
# print("写入中。。。")
# for m in range(2, max_row + 1 - 2160):
#     n = 101  # chr(97)='a'
#     o = 108
#     n = chr(n)  # ASCII字符
#     o = chr(o)
#     i = '%s%d' % (o, m + 2160)  # 读取单元格编号
#     j = '%s%d' % (n, m)  # 写入单元格编号
#     cell1 = sheet1[i].value  # 获取data单元格数据
#     sheet2[j].value = cell1  # 赋值到test单元格
#
# # 填写1-1到2-28数据
# for m in range(6602, 8018):
#     n = 101  # chr(97)='a'
#     o = 108
#     n = chr(n)  # ASCII字符
#     o = chr(o)
#     i = '%s%d' % (o, m - 6600)  # 读取单元格编号
#     j = '%s%d' % (n, m)  # 写入单元格编号
#     cell1 = sheet1[i].value  # 获取data单元格数据
#     sheet2[j].value = cell1  # 赋值到test单元格
#
# # 填写2-29数据
# for m in range(8018, 8042):
#     n = 101  # chr(97)='a'
#     o = 108
#     n = chr(n)  # ASCII字符
#     o = chr(o)
#     i = '%s%d' % (o, m - 6624)  # 读取单元格编号
#     j = '%s%d' % (n, m)  # 写入单元格编号
#     cell1 = sheet1[i].value  # 获取data单元格数据
#     sheet2[j].value = cell1  # 赋值到test单元格
#
# # 填写3-1到3-31数据
# for m in range(8042, max_row + 1):
#     n = 101  # chr(97)='a'
#     o = 108
#     n = chr(n)  # ASCII字符
#     o = chr(o)
#     i = '%s%d' % (o, m - 6624)  # 读取单元格编号
#     j = '%s%d' % (n, m)  # 写入单元格编号
#     cell1 = sheet1[i].value  # 获取data单元格数据
#     sheet2[j].value = cell1  # 赋值到test单元格


# 完成收尾工作
wb3.save(resultFileName + '.xlsx')  # 保存数据
print("写入完成")
os.remove('somethingYouNeed.xlsx')
os.remove('somethingYouNeedToo.csv')
os.remove('somethingYouNeedToo.xlsx')
wb1.close()  # 关闭excel
wb3.close()
